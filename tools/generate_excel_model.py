"""
GTT Center Perth -- Excel Financial Model Generator (Phase 10, Document Generation).

Generates outputs/GTT-Center-Perth-Financial-Model.xlsx from the canonical
data/model layer, via tools/document_data.py. This script does NOT compute
or invent any new figure -- every number placed in a cell traces to
tools/master_financial_model.py's live output or a data/canonical/*.yml
record, and cross-referencing formulas are used within the workbook wherever
practical so totals are linked, not duplicated hard-coded values.

Hard scope boundary (unchanged from the phase brief): does not choose a
primary scenario, does not invent an opening cash balance or financing
structure, does not present the bounded A$357,390-577,180 funding range as a
single exact figure, does not let startup/capex costs enter the operating
P&L sheet.

Usage:
    python tools/generate_excel_model.py
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import document_data as dd  # noqa: E402

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

REPO_ROOT = Path(__file__).resolve().parent.parent
OUTPUT_PATH = REPO_ROOT / "outputs" / "GTT-Center-Perth-Financial-Model.xlsx"

# ---------------------------------------------------------------------------
# Styling constants -- applied consistently across every sheet.
# ---------------------------------------------------------------------------
NAVY = "1F3864"
LIGHT_BLUE = "DCE6F1"
AMBER = "FFF2CC"
RED_TEXT = "C00000"
GREY = "F2F2F2"
WHITE = "FFFFFF"

TITLE_FONT = Font(name="Calibri", size=16, bold=True, color=WHITE)
SECTION_FONT = Font(name="Calibri", size=13, bold=True, color=WHITE)
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color=WHITE)
LABEL_FONT = Font(name="Calibri", size=10, bold=True)
BODY_FONT = Font(name="Calibri", size=10)
ITALIC_FONT = Font(name="Calibri", size=9, italic=True, color="595959")
WARNING_FONT = Font(name="Calibri", size=10, bold=True, color=RED_TEXT)

TITLE_FILL = PatternFill("solid", fgColor=NAVY)
SECTION_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
ALT_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
WARNING_FILL = PatternFill("solid", fgColor=AMBER)

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CURRENCY_FMT = '#,##0.00;[RED]-#,##0.00'
CURRENCY_WHOLE_FMT = '#,##0;[RED]-#,##0'
PCT_FMT = '0.00"%"'
NUM_FMT = '0.000'


def style_title(ws: Worksheet, row: int, text: str, span: int = 8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 28


def style_section(ws: Worksheet, row: int, text: str, span: int = 8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 20


def style_header_row(ws: Worksheet, row: int, headers: list):
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_row(ws: Worksheet, row: int, values: list, fmts: dict = None, alt=False, bold_col1=False):
    fmts = fmts or {}
    for i, v in enumerate(values, start=1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.border = BORDER
        cell.font = LABEL_FONT if (bold_col1 and i == 1) else BODY_FONT
        if alt:
            cell.fill = ALT_FILL
        if i in fmts:
            cell.number_format = fmts[i]
    return row


def note_block(ws: Worksheet, row: int, lines: list, span: int = 8, warning=False):
    for line in lines:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        cell = ws.cell(row=row, column=1, value=line)
        cell.font = WARNING_FONT if warning else ITALIC_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        if warning:
            cell.fill = WARNING_FILL
        ws.row_dimensions[row].height = 15
        row += 1
    return row


def set_widths(ws: Worksheet, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


# ---------------------------------------------------------------------------
# Sheet 1 -- Executive Summary
# ---------------------------------------------------------------------------
def build_executive_summary(wb, data: dd.DocumentData):
    ws = wb.active
    ws.title = "1. Executive Summary"
    set_widths(ws, {1: 34, 2: 22, 3: 22, 4: 4, 5: 30})

    style_title(ws, 1, "GTT Center Perth — Master Financial Model")
    r = 2
    ws.cell(row=r, column=1, value="Executive Summary — Table 1 vs Table 2").font = Font(size=12, italic=True)
    r += 2

    r = note_block(ws, r, [
        "IMPORTANT — READ FIRST",
        "This workbook presents CANONICAL (validated, deterministic) figures alongside HISTORICAL/INHERITED,",
        "MODELLED, and BOUNDED figures. Every figure is labelled with its governance status. Neither Table 1",
        "(18 clients/day) nor Table 2 (12 clients/day) is chosen as the venture's primary scenario — see",
        "VERIFICATION-TRACKER.md item 1m, still open. The funding requirement (Section 7) is a BOUNDED RANGE,",
        "not an exact figure — do not quote a single number from it. No opening cash balance or financing",
        "structure has been assumed anywhere in this workbook. Full source: docs/architecture/ methodology docs.",
    ], warning=True)
    r += 1

    style_section(ws, r, "Scenario Comparison — Month 5+ Steady State"); r += 1
    headers = ["Metric", "Table 1 (18/day)", "Table 2 (12/day)", "", "Status"]
    style_header_row(ws, r, headers); r += 1
    sc = data.scenario_comparison
    rows = [
        ("Client volume (per day)", sc["scenario_table_1"]["client_volume_per_day"], sc["scenario_table_2"]["client_volume_per_day"], "", "VERIFIED"),
        ("Monthly revenue (AUD)", sc["scenario_table_1"]["steady_state_revenue"], sc["scenario_table_2"]["steady_state_revenue"], "", "CANONICAL / CALCULATED"),
        ("Monthly payroll, incl. super (AUD)", sc["scenario_table_1"]["steady_state_payroll"], sc["scenario_table_2"]["steady_state_payroll"], "", "CALCULATED"),
        ("Monthly operating expenses (AUD)", sc["scenario_table_1"]["steady_state_opex"], sc["scenario_table_2"]["steady_state_opex"], "", "CALCULATED"),
        ("Total operating costs (AUD)", sc["scenario_table_1"]["steady_state_total_operating_costs"], sc["scenario_table_2"]["steady_state_total_operating_costs"], "", "CALCULATED"),
        ("Net operating result (AUD/month)", sc["scenario_table_1"]["steady_state_net_operating_result"], sc["scenario_table_2"]["steady_state_net_operating_result"], "", "CALCULATED"),
        ("Break-even AM client volume/day", data.breakeven["scenario_table_1"]["breakeven_am_client_volume_per_day"], data.breakeven["scenario_table_2"]["breakeven_am_client_volume_per_day"], "", "CALCULATED, bounded scope"),
        ("Margin of safety (clients/day)", data.breakeven["scenario_table_1"]["margin_of_safety_clients_per_day"], data.breakeven["scenario_table_2"]["margin_of_safety_clients_per_day"], "", "CALCULATED"),
        ("24-month net operating result (AUD)", data.totals_24mo["scenario_table_1"]["total_net_operating_result_24mo"], data.totals_24mo["scenario_table_2"]["total_net_operating_result_24mo"], "", "CALCULATED"),
        ("Annualised revenue (AUD)", sc["scenario_table_1"]["annualised_revenue"], sc["scenario_table_2"]["annualised_revenue"], "", "CALCULATED"),
        ("Annualised net operating result (AUD)", sc["scenario_table_1"]["annualised_net_operating_result"], sc["scenario_table_2"]["annualised_net_operating_result"], "", "CALCULATED"),
        ("Operating cash trough (AUD)", -abs(data.cash_flow["scenario_table_1"]["trough_cumulative_position"]), -abs(data.cash_flow["scenario_table_2"]["trough_cumulative_position"]), "", "CALCULATED, operating-cash-only"),
    ]
    for i, (label, v1, v2, _, status) in enumerate(rows):
        alt = i % 2 == 0
        write_row(ws, r, [label, v1, v2, "", status], fmts={2: CURRENCY_FMT, 3: CURRENCY_FMT}, alt=alt, bold_col1=True)
        r += 1
    r += 1

    style_section(ws, r, "Bounded Funding Requirement"); r += 1
    fr = data.funding
    combined = fr["combined_funding_requirement_bounded"]["primary_method"]
    write_row(ws, r, ["Pre-opening capital (AUD)", fr["pre_opening_capital"]["range_low"], fr["pre_opening_capital"]["range_high"], "", "CALCULATED, universal (not scenario-specific)"],
              fmts={2: CURRENCY_FMT, 3: CURRENCY_FMT}, bold_col1=True); r += 1
    write_row(ws, r, ["Working capital reserve, historical basis (AUD)", fr["opening_working_capital"]["historical_reserve_method"]["range_low"], fr["opening_working_capital"]["historical_reserve_method"]["range_high"], "", "MODELLED, stale basis"],
              fmts={2: CURRENCY_FMT, 3: CURRENCY_FMT}, alt=True, bold_col1=True); r += 1
    write_row(ws, r, ["BOUNDED FUNDING REQUIREMENT — RANGE, NOT EXACT (AUD)", combined["range_low"], combined["range_high"], "", "CALCULATED — BOUNDED, not exact"],
              fmts={2: CURRENCY_FMT, 3: CURRENCY_FMT}, bold_col1=True); r += 2
    r = note_block(ws, r, [
        "The funding requirement above is a RANGE (A$357,390 – A$577,180), not a single point estimate. The",
        "underlying 6–9 historical startup-capital ranges remain unreconciled (docs/architecture/",
        "STARTUP-COST-RECONCILIATION.md) — see the Funding Requirement sheet for full detail.",
    ])
    r += 1

    style_section(ws, r, "Key Risks / Unresolved Assumptions (summary — see Sheet 9 for full list)"); r += 1
    risks = [
        "Historical revenue/cost figures differ from canonical (A$2,576.36/month gap, origin permanently unresolved — item 36).",
        "10% PM pre-booking discount is NOT applied anywhere in this model (item 39) — a revenue-overstatement risk.",
        "AM labour is modelled fixed from Month 1 — a conservative, disclosed simplification, not a discovered fact (item 43).",
        "Superannuation is applied to most, but not all, payroll components — Opening-Time Increment/Receptionist-Relief remain unresolved.",
        "Consumables/laundry FIXED-vs-VARIABLE classification remains unresolved — no 'total visits' definition exists (item 45).",
        "Funding requirement is bounded, not exact — see Section 7.",
    ]
    for risk in risks:
        write_row(ws, r, [risk], bold_col1=False)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 1

    ws.freeze_panes = "A5"
    return ws


# ---------------------------------------------------------------------------
# Sheet 2 -- Revenue Model
# ---------------------------------------------------------------------------
def build_revenue_model(wb, data: dd.DocumentData):
    ws = wb.create_sheet("2. Revenue Model")
    set_widths(ws, {1: 26, 2: 18, 3: 18, 4: 18, 5: 18, 6: 18, 7: 18})

    style_title(ws, 1, "Revenue Model")
    r = 3
    note_lines = [
        "Canonical formula (docs/architecture/CANONICAL-REVENUE-METHODOLOGY.md):",
        "Monthly Revenue = (clients/day × AM price × operating weekdays) + (clients/day × AM price × operating Saturdays)",
        "                 + (PM weekday sessions × PM price × operating weekdays) + (PM Saturday sessions × PM price × operating Saturdays) + Ancillary",
        "Canonical inputs: pricing.yml#am_price_used_for_revenue (A$250), pricing.yml#pm_alacarte_average (A$95),",
        "client_assumptions.yml#operating_days_per_month_weekday (22), #operating_saturdays_per_month (4.33),",
        "client_assumptions.yml#pm_steady_state_capacity (16), revenue_assumptions.yml#rev_pm_saturday_sessions (8).",
    ]
    r = note_block(ws, r, note_lines, span=7)
    r += 1

    for sid in dd.SCENARIOS:
        style_section(ws, r, f"{data.label(sid)} — Month 1-5+ Revenue Ramp", span=7); r += 1
        headers = ["Month", "AM Revenue", "PM Revenue", "Ancillary Revenue", "Total Revenue", "% of Steady State", "Status"]
        style_header_row(ws, r, headers); r += 1
        for m in [1, 2, 3, 4, 5]:
            pnl = data.pnl_24mo[sid][m - 1]
            pct = round(pnl["revenue"]["total_revenue"] / data.steady_state[sid]["revenue"]["total_revenue"] * 100, 2)
            write_row(
                ws, r,
                [f"M{m}" if m < 5 else "M5+ (steady state)", pnl["revenue"]["am_revenue"], pnl["revenue"]["pm_revenue"],
                 pnl["revenue"]["ancillary_revenue"], pnl["revenue"]["total_revenue"], pct, "CALCULATED"],
                fmts={2: CURRENCY_FMT, 3: CURRENCY_FMT, 4: CURRENCY_FMT, 5: CURRENCY_FMT, 6: PCT_FMT},
                alt=(m % 2 == 0), bold_col1=True,
            )
            r += 1
        r += 1

    r = note_block(ws, r, [
        "Months 6-24: revenue held FLAT at Month 5+ steady state — no growth invented past Month 5,",
        "per docs/architecture/CANONICAL-REVENUE-METHODOLOGY.md and explicit instruction. See Sheet 4 (P&L) for the full 24-month series.",
        "Ancillary revenue is A$0.00 throughout — EXCLUDED from the baseline per Anthony's 2026-07-30 instruction (VERIFIED), shown explicitly not omitted.",
    ], span=7)
    ws.freeze_panes = "A4"
    return ws


# ---------------------------------------------------------------------------
# Sheet 3 -- Cost Model
# ---------------------------------------------------------------------------
def build_cost_model(wb, data: dd.DocumentData):
    ws = wb.create_sheet("3. Cost Model")
    set_widths(ws, {1: 32, 2: 18, 3: 18, 4: 4, 5: 32, 6: 18, 7: 18})

    style_title(ws, 1, "Cost Model")
    r = 3
    r = note_block(ws, r, [
        "Costs = Fixed (Non-Wage Overhead excl. marketing) + Variable (marketing ramp) + Payroll (incl. superannuation).",
        "Source: data/canonical/cost_ramp.yml, docs/architecture/COST-RAMP-METHODOLOGY.md.",
    ], span=7)
    r += 1

    style_section(ws, r, "Payroll Breakdown — Month 5+ Steady State", span=7); r += 1
    headers = ["Payroll Component", "Table 1 (AUD)", "Table 2 (AUD)", "", "Superannuation Treatment", "", ""]
    style_header_row(ws, r, headers); r += 1
    t1_bd = data.steady_state["scenario_table_1"]["payroll_breakdown"]
    t2_bd = data.steady_state["scenario_table_2"]["payroll_breakdown"]
    payroll_rows = [
        ("AM weekday — treatment staff (8 staff)", t1_bd["am_weekday_treatment_staff"], t2_bd["am_weekday_treatment_staff"], "Already includes super (source table's own 'incl. super' label) — NOT added again"),
        ("AM weekday — phlebotomist (2 staff)", t1_bd["am_weekday_phlebotomist"], t2_bd["am_weekday_phlebotomist"], "Super ADDED (12%) — source table's Phlebotomist row is NOT labelled 'incl. super'"),
        ("AM Saturday direct labour", t1_bd["am_saturday_direct_labor"], t2_bd["am_saturday_direct_labor"], "Super ADDED (12%) — hourly-rate-based, exclusive of super"),
        ("PM weekday direct labour", t1_bd["pm_weekday_direct_labor"], t2_bd["pm_weekday_direct_labor"], "Super ADDED (12%) — hourly-rate-based"),
        ("PM Saturday direct labour", t1_bd["pm_saturday_direct_labor"], t2_bd["pm_saturday_direct_labor"], "Super ADDED (12%) — hourly-rate-based"),
        ("Opening-time increment", t1_bd["opening_time_increment"], t2_bd["opening_time_increment"], "NOT super-eligible-confirmed — UNRESOLVED (conflict_superannuation_partial_coverage)"),
        ("Receptionist / Relief Pool", t1_bd["receptionist_relief"], t2_bd["receptionist_relief"], "NOT super-eligible-confirmed — UNRESOLVED (conflict_superannuation_partial_coverage)"),
        ("Superannuation (12% OTE, applied per above)", t1_bd["superannuation"], t2_bd["superannuation"], "See docs/architecture/COST-RAMP-METHODOLOGY.md §4a"),
        ("Workers Compensation (1.7%)", t1_bd["workers_comp"], t2_bd["workers_comp"], "MODELLED, % of Direct Labor + Opening Costs"),
    ]
    for i, (label, v1, v2, note) in enumerate(payroll_rows):
        write_row(ws, r, [label, v1, v2, "", note], fmts={2: CURRENCY_FMT, 3: CURRENCY_FMT}, alt=(i % 2 == 0), bold_col1=True)
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
        r += 1
    # Total row -- FORMULA, summing the cells above (not a duplicated hard-coded value).
    total_row = r
    first_data_row = total_row - len(payroll_rows)
    write_row(ws, r, ["TOTAL PAYROLL (incl. super)", f"=SUM(B{first_data_row}:B{total_row-1})", f"=SUM(C{first_data_row}:C{total_row-1})", "", "= sum of all rows above (linked formula)"],
              fmts={2: CURRENCY_FMT, 3: CURRENCY_FMT}, bold_col1=True)
    for c in range(1, 4):
        ws.cell(row=r, column=c).font = Font(bold=True)
    r += 2

    style_section(ws, r, "Operating Expenses — Fixed vs. Variable", span=7); r += 1
    headers = ["Component", "Table 1 = Table 2 (AUD)", "", "", "Classification / Status", "", ""]
    style_header_row(ws, r, headers); r += 1
    fixed = data.steady_state["scenario_table_1"]["operating_expenses_breakdown"]["fixed_costs"]
    variable = data.steady_state["scenario_table_1"]["operating_expenses_breakdown"]["variable_costs"]
    write_row(ws, r, ["Fixed Non-Wage Overhead (excl. marketing)", fixed, "", "", "FIXED — status-quo classification (universal, both scenarios)"], fmts={2: CURRENCY_FMT}, bold_col1=True); r += 1
    write_row(ws, r, ["Marketing (variable, ramps Month 1-4)", variable, "", "", "SEMI_VARIABLE — the only documented ramped opex line"], fmts={2: CURRENCY_FMT}, alt=True, bold_col1=True); r += 1
    write_row(ws, r, ["TOTAL OPERATING EXPENSES", "=B" + str(r - 2) + "+B" + str(r - 1), "", "", "= Fixed + Variable (linked formula)"], fmts={2: CURRENCY_FMT}, bold_col1=True)
    for c in range(1, 3):
        ws.cell(row=r, column=c).font = Font(bold=True)
    r += 2

    r = note_block(ws, r, [
        "UNRESOLVED cost classifications (not resolved by this workbook):",
        "GTT supplies / general consumables / laundry — FIXED per profit-loss-tables.md vs. VARIABLE per unit-economics.md (historical) — conflict_variable_vs_fixed_classification_carried_forward.",
        "Insurance — modelled A$400/month vs. itemised A$975-1,583/month (mandatory policies only) — conflict_insurance_estimate. See Sheet 8, Sensitivity.",
        "AM labour ramp — modelled FIXED from Month 1 (conservative, disclosed simplification) — conflict_am_labor_ramp_unmodelled.",
    ], span=7, warning=True)
    ws.freeze_panes = "A4"
    return ws


def _pnl_block(ws, r, sid, data: "dd.DocumentData"):
    """Writes one scenario's 24-month P&L block with LIVE FORMULAS: Gross
    Contribution, Total Operating Costs, Net Operating Result, and Cumulative
    are all computed by in-sheet formulas referencing the Revenue/Payroll/
    OpEx input cells -- not duplicated hard-coded values."""
    style_section(ws, r, f"{data.label(sid)} — 24-Month P&L", span=8); r += 1
    headers = ["Month", "Revenue", "Payroll", "Gross Contribution", "Operating Expenses", "Total Operating Costs", "Net Operating Result", "Cumulative"]
    style_header_row(ws, r, headers); r += 1
    first_row = r
    for i, pnl in enumerate(data.pnl_24mo[sid], start=1):
        rev = pnl["revenue"]["total_revenue"]
        payroll = pnl["payroll"]
        opex = pnl["operating_expenses"]
        row_vals = [
            f"M{i}", rev, payroll,
            f"=B{r}-C{r}",                              # Gross Contribution = Revenue - Payroll
            opex,
            f"=C{r}+E{r}",                               # Total Operating Costs = Payroll + OpEx
            f"=B{r}-F{r}",                               # Net Operating Result = Revenue - Total Operating Costs
            f"=G{r}" if i == 1 else f"=H{r-1}+G{r}",     # Cumulative
        ]
        write_row(
            ws, r, row_vals,
            fmts={2: CURRENCY_FMT, 3: CURRENCY_FMT, 4: CURRENCY_FMT, 5: CURRENCY_FMT, 6: CURRENCY_FMT, 7: CURRENCY_FMT, 8: CURRENCY_FMT},
            alt=(i % 2 == 0), bold_col1=True,
        )
        r += 1
    write_row(
        ws, r,
        ["24-MONTH TOTAL", f"=SUM(B{first_row}:B{r-1})", f"=SUM(C{first_row}:C{r-1})", f"=SUM(D{first_row}:D{r-1})",
         f"=SUM(E{first_row}:E{r-1})", f"=SUM(F{first_row}:F{r-1})", f"=SUM(G{first_row}:G{r-1})", f"=H{r-1}"],
        fmts={2: CURRENCY_FMT, 3: CURRENCY_FMT, 4: CURRENCY_FMT, 5: CURRENCY_FMT, 6: CURRENCY_FMT, 7: CURRENCY_FMT, 8: CURRENCY_FMT},
        bold_col1=True,
    )
    for c in range(1, 9):
        ws.cell(row=r, column=c).font = Font(bold=True)
        ws.cell(row=r, column=c).fill = ALT_FILL
    r += 2
    return r


# ---------------------------------------------------------------------------
# Sheet 4 -- P&L / Operating Model (24-month, both scenarios, live formulas)
# ---------------------------------------------------------------------------
def build_pnl_model(wb, data: "dd.DocumentData"):
    ws = wb.create_sheet("4. P&L Operating Model")
    set_widths(ws, {1: 12, 2: 16, 3: 16, 4: 18, 5: 18, 6: 20, 7: 18, 8: 18})
    style_title(ws, 1, "P&L / Operating Model — Month 1-24", span=8)
    r = 3
    r = note_block(ws, r, [
        "Gross Contribution, Total Operating Costs, Net Operating Result, and Cumulative columns are LIVE FORMULAS",
        "referencing the Revenue/Payroll/Operating Expenses cells in this sheet — not duplicated hard-coded values.",
        "Months 6-24 repeat Month 5's steady-state Revenue/Payroll/OpEx inputs exactly (no growth invented past Month 5).",
        "Startup costs and capex are NOT included anywhere in this sheet — operating P&L only.",
    ], span=8)
    r += 1
    for sid in dd.SCENARIOS:
        r = _pnl_block(ws, r, sid, data)
    ws.freeze_panes = "A4"
    return ws


# ---------------------------------------------------------------------------
# Sheet 5 -- Cash Flow (operating-cash basis only)
# ---------------------------------------------------------------------------
def build_cash_flow(wb, data: "dd.DocumentData"):
    ws = wb.create_sheet("5. Cash Flow")
    set_widths(ws, {1: 12, 2: 24, 3: 22, 4: 40})
    style_title(ws, 1, "Cash Flow — Operating Basis Only", span=4)
    r = 3
    r = note_block(ws, r, [
        "IMPORTANT: this is an OPERATING-CASH-basis view only (net operating result used as an accrual-basis proxy",
        "for cash — NOT a true cash-basis forecast, since no debtor/creditor timing data exists anywhere in this repo).",
        "This is NOT a complete funding schedule — it EXCLUDES startup capital deployment, capex, and any opening",
        "cash balance (none assumed — see Sheet 7, Funding Requirement, for the separate, bounded funding range).",
    ], span=4, warning=True)
    r += 1

    for sid in dd.SCENARIOS:
        style_section(ws, r, f"{data.label(sid)} — Monthly & Cumulative Operating Cash Position", span=4); r += 1
        headers = ["Month", "Net Operating Result (AUD)", "Cumulative Position (AUD)", "Status"]
        style_header_row(ws, r, headers); r += 1
        cf = data.cash_flow[sid]
        for row in cf["rows"]:
            write_row(
                ws, r, [f"M{row['forecast_month']}", row["net_monthly_cash_movement"], row["cumulative_position"], "CALCULATED"],
                fmts={2: CURRENCY_FMT, 3: CURRENCY_FMT}, alt=(row["forecast_month"] % 2 == 0), bold_col1=True,
            )
            r += 1
        write_row(ws, r, [
            f"Operating cash trough: A${cf['trough_cumulative_position']:,.2f} at Month {cf['trough_month']}",
            "", "", "OPERATING-CASH-ONLY — NOT the total funding requirement (see Sheet 7)",
        ], bold_col1=True)
        for c in range(1, 5):
            ws.cell(row=r, column=c).font = WARNING_FONT
            ws.cell(row=r, column=c).fill = WARNING_FILL
        r += 2
    ws.freeze_panes = "A4"
    return ws


# ---------------------------------------------------------------------------
# Sheet 6 -- Break-Even
# ---------------------------------------------------------------------------
def build_breakeven(wb, data: "dd.DocumentData"):
    ws = wb.create_sheet("6. Break-Even")
    set_widths(ws, {1: 38, 2: 22, 3: 22})
    style_title(ws, 1, "Break-Even Analysis", span=3)
    r = 3
    r = note_block(ws, r, [
        "A traditional contribution-margin break-even (fixed costs / contribution margin %) is NOT computed.",
        "Reason: nearly every cost in this model is classified FIXED (including AM/PM payroll under the current,",
        "disclosed treatment), and the GTT-supplies/consumables/laundry classification remains unresolved — a",
        "reliable fixed-vs-variable cost split does not exist in the canonical cost data.",
        "What IS computed: an AM client-volume break-even, holding Month 5+ PM revenue/payroll/opex fixed and",
        "treating AM revenue as the one genuinely linear-in-volume component (price x operating days).",
    ], span=3, warning=True)
    r += 1
    headers = ["Metric", "Table 1 (18/day)", "Table 2 (12/day)"]
    style_header_row(ws, r, headers); r += 1
    be1, be2 = data.breakeven["scenario_table_1"], data.breakeven["scenario_table_2"]
    rows = [
        ("Break-even AM client volume/day", be1["breakeven_am_client_volume_per_day"], be2["breakeven_am_client_volume_per_day"], NUM_FMT),
        ("Break-even monthly revenue (AUD)", be1["breakeven_monthly_revenue"], be2["breakeven_monthly_revenue"], CURRENCY_FMT),
        ("Committed client volume/day", be1["committed_client_volume_per_day"], be2["committed_client_volume_per_day"], NUM_FMT),
        ("Margin of safety (clients/day)", be1["margin_of_safety_clients_per_day"], be2["margin_of_safety_clients_per_day"], NUM_FMT),
    ]
    for i, (label, v1, v2, fmt) in enumerate(rows):
        write_row(ws, r, [label, v1, v2], fmts={2: fmt, 3: fmt}, alt=(i % 2 == 0), bold_col1=True)
        r += 1
    r += 1
    r = note_block(ws, r, [
        "Basis: Month 5+ (steady state) costs and PM/ancillary revenue held fixed; AM revenue treated as the only",
        "volume-linear component. See data/models/master_financial_model.yml#compute_breakeven for full method.",
    ], span=3)
    ws.freeze_panes = "A4"
    return ws


# ---------------------------------------------------------------------------
# Sheet 7 -- Funding Requirement (BOUNDED, not exact)
# ---------------------------------------------------------------------------
def build_funding_requirement(wb, data: "dd.DocumentData"):
    ws = wb.create_sheet("7. Funding Requirement")
    set_widths(ws, {1: 46, 2: 18, 3: 18, 4: 46})
    style_title(ws, 1, "Funding Requirement — BOUNDED, NOT EXACT", span=4)
    r = 3
    r = note_block(ws, r, [
        "docs/VERIFICATION-TRACKER.md item 47 — investigated and BOUNDED (OUTCOME 2), NOT resolved to an exact figure.",
        "The underlying 6-9-range startup-cost reconciliation (docs/architecture/STARTUP-COST-RECONCILIATION.md) remains",
        "genuinely unresolved. DO NOT quote a single number from this sheet — always quote the range.",
    ], span=4, warning=True)
    r += 1
    fr = data.funding

    style_section(ws, r, "(A) Pre-Opening Capital — universal, not scenario-specific", span=4); r += 1
    poc = fr["pre_opening_capital"]
    write_row(ws, r, ["Pre-opening capital range (AUD)", poc["range_low"], poc["range_high"],
                       "CALCULATED — construction, equipment/furniture/signage, legal/lease bond; excludes working capital"],
              fmts={2: CURRENCY_FMT, 3: CURRENCY_FMT}, bold_col1=True); r += 1
    write_row(ws, r, ["Timing", "", "", poc["timing_classification"]], bold_col1=True); r += 2

    style_section(ws, r, "(B) Opening Working Capital — two disclosed methods, neither chosen", span=4); r += 1
    hist = fr["opening_working_capital"]["historical_reserve_method"]
    write_row(ws, r, ["Historical reserve method (AUD)", hist["range_low"], hist["range_high"],
                       "MODELLED, stale basis (pre-rebase Month 1-3 loss estimate, item 30)"],
              fmts={2: CURRENCY_FMT, 3: CURRENCY_FMT}, bold_col1=True); r += 1
    for cc in fr["opening_working_capital"]["operating_cash_trough_cross_check"]["results"]:
        label = f"Operating-cash-trough cross-check — {dd.SCENARIO_LABELS[cc['scenario_id']]}"
        write_row(ws, r, [label, cc["trough_value"], "", f"Alternative cross-check only, trough Month {cc['trough_month']} — NOT chosen over the historical reserve"],
                  fmts={2: CURRENCY_FMT}, alt=True, bold_col1=True); r += 1
    r += 1

    style_section(ws, r, "(C) Combined Bounded Funding Requirement", span=4); r += 1
    combined = fr["combined_funding_requirement_bounded"]["primary_method"]
    write_row(ws, r, ["PRIMARY METHOD: (A) + historical Working Capital Reserve", combined["range_low"], combined["range_high"],
                       "Exact match to startup_costs.yml#total_current_state_component_sum — no new figure invented"],
              fmts={2: CURRENCY_FMT, 3: CURRENCY_FMT}, bold_col1=True)
    for c in (1, 2, 3, 4):
        ws.cell(row=r, column=c).font = Font(bold=True, size=11)
        ws.cell(row=r, column=c).fill = ALT_FILL
    r += 1
    for alt in fr["combined_funding_requirement_bounded"]["alternative_cross_check_method"]["results"]:
        label = f"Alternative: (A) + operating cash trough — {dd.SCENARIO_LABELS[alt['scenario_id']]}"
        write_row(ws, r, [label, alt["range_low"], alt["range_high"], "Illustrative alternative only, not chosen as more correct"],
                  fmts={2: CURRENCY_FMT, 3: CURRENCY_FMT}, alt=True, bold_col1=True); r += 1
    r += 1

    r = note_block(ws, r, [
        "Anthony's own adopted, reconciled total (docs/CURRENT-STATE.md §7.4): A$292,335 - A$594,900 — close to but",
        "NOT identical to the primary method range above (CURRENT-STATE.md itself discloses this gap). Both figures",
        "remain valid, disclosed, side by side — neither superseded by this workbook.",
        "",
        "NOT INCLUDED anywhere in this sheet: an opening cash balance (none exists in this repo), a financing/debt/",
        "equity structure (none assumed), or the operating cash trough summed ON TOP of the working capital reserve",
        "(would double-count the same 'survive the ramp' concept — see deliberate_non_assumptions in the model YAML).",
    ], span=4)
    ws.freeze_panes = "A4"
    return ws


# ---------------------------------------------------------------------------
# Sheet 8 -- Sensitivity Analysis (existing modelled sensitivities only)
# ---------------------------------------------------------------------------
def build_sensitivity(wb, data: "dd.DocumentData"):
    ws = wb.create_sheet("8. Sensitivity Analysis")
    set_widths(ws, {1: 26, 2: 18, 3: 18, 4: 18, 5: 18, 6: 40})
    style_title(ws, 1, "Sensitivity Analysis", span=6)
    r = 3
    r = note_block(ws, r, [
        "Only the two sensitivities already computed by tools/master_financial_model.py are shown here — no new",
        "sensitivity is invented for this workbook.",
    ], span=6)
    r += 1
    for sid in dd.SCENARIOS:
        style_section(ws, r, f"{data.label(sid)} — Client Volume Sensitivity (50/75/100/125% of committed)", span=6); r += 1
        headers = ["% of Committed Volume", "Client Volume/Day", "AM Revenue", "Total Revenue", "Net Operating Result", "Note"]
        style_header_row(ws, r, headers); r += 1
        for i, row in enumerate(data.sensitivity_client_volume[sid]):
            write_row(ws, r, [f"{row['pct_of_committed']}%", row["client_volume_per_day"], row["am_revenue"], row["total_revenue"], row["net_operating_result"], row["note"]],
                      fmts={2: NUM_FMT, 3: CURRENCY_FMT, 4: CURRENCY_FMT, 5: CURRENCY_FMT}, alt=(i % 2 == 0), bold_col1=True)
            ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=6)
            r += 1
        r += 1

    for sid in dd.SCENARIOS:
        style_section(ws, r, f"{data.label(sid)} — Insurance Sensitivity (modelled vs. itemised, conflict_insurance_estimate)", span=6); r += 1
        headers = ["Insurance Scenario", "Monthly Insurance", "Delta vs. Modelled", "Adjusted Total Op. Costs", "Adjusted Net Op. Result", ""]
        style_header_row(ws, r, headers); r += 1
        for i, row in enumerate(data.sensitivity_insurance[sid]):
            write_row(ws, r, [row["scenario"], row["monthly_insurance"], row["delta_vs_modelled"], row["adjusted_total_operating_costs"], row["adjusted_net_operating_result"], ""],
                      fmts={2: CURRENCY_FMT, 3: CURRENCY_FMT, 4: CURRENCY_FMT, 5: CURRENCY_FMT}, alt=(i % 2 == 0), bold_col1=True)
            r += 1
        r += 1
    ws.freeze_panes = "A4"
    return ws


# ---------------------------------------------------------------------------
# Sheet 9 -- Assumptions & Open Issues
# ---------------------------------------------------------------------------
def build_assumptions_open_issues(wb, data: "dd.DocumentData"):
    ws = wb.create_sheet("9. Assumptions & Open Issues")
    set_widths(ws, {1: 10, 2: 40, 3: 18, 4: 55})
    style_title(ws, 1, "Assumptions & Open Issues", span=4)
    r = 3
    r = note_block(ws, r, [
        "This sheet reproduces the currently-unresolved items from docs/VERIFICATION-TRACKER.md most relevant to the",
        "figures presented in this workbook. It does not resolve any of them — presentation only, per this phase's scope.",
    ], span=4)
    r += 1
    headers = ["Tracker Item", "Description", "Status", "Notes"]
    style_header_row(ws, r, headers); r += 1
    open_issues = [
        ("30", "Working capital reserve (A$85,000-110,000) basis is a pre-2026-08-05-rebase Month 1-3 loss estimate", "UNRESOLVED (stale basis)", "See Sheet 7 (B) Historical reserve method"),
        ("36", "Historical Table 1/2 monthly revenue (A$157,792.16 / A$118,297.16) vs. canonical (A$155,215.80 / A$115,720.80) — A$2,576.36 gap", "RESOLVED (nuanced) — historical origin unresolved, future methodology resolved", "See Sheet 10 and docs/architecture/MASTER-FINANCIAL-MODEL-METHODOLOGY.md"),
        ("39", "PM discount / pricing basis for the a-la-carte average used in PM revenue", "UNRESOLVED", "See data/canonical/pricing.yml#conflicts"),
        ("41-42", "Revenue-ramp origin — Month 1-4 ramp percentages' original derivation", "UNRESOLVED origin, canonical going forward", "See data/canonical/revenue_ramp.yml#conflicts"),
        ("43", "AM labour ramp not modelled — payroll held FIXED from Month 1 rather than flexed with the ramp", "UNRESOLVED, disclosed conservative simplification", "conflict_am_labor_ramp_unmodelled, see Sheet 3 and Sheet 8"),
        ("45", "GTT supplies / consumables / laundry — FIXED vs. VARIABLE classification conflict", "UNRESOLVED", "conflict_variable_vs_fixed_classification_carried_forward, see Sheet 3"),
        ("46", "Superannuation — RESOLVED for 6 of 8 payroll components; Opening-Time-Increment and Receptionist/Relief Pool remain super-uncertain", "PARTIALLY RESOLVED", "conflict_superannuation_partial_coverage, see Sheet 3"),
        ("47", "Funding requirement — bounded range only (A$357,390-577,180), not an exact figure; underlying 6-9-range startup-cost reconciliation unresolved", "PARTIALLY RESOLVED (bounded)", "See Sheet 7 and docs/architecture/FUNDING-REQUIREMENT-INVESTIGATION.md"),
        ("n/a", "Insurance — modelled A$400/month vs. itemised A$975-1,583/month (mandatory policies only)", "UNRESOLVED", "conflict_insurance_estimate, see Sheet 8 Sensitivity"),
    ]
    for i, (item, desc, status, notes) in enumerate(open_issues):
        write_row(ws, r, [item, desc, status, notes], alt=(i % 2 == 0), bold_col1=True)
        ws.row_dimensions[r].height = 30
        for c in range(1, 5):
            ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    ws.freeze_panes = "A4"
    return ws


# ---------------------------------------------------------------------------
# Sheet 10 -- Source & Methodology Register
# ---------------------------------------------------------------------------
def build_source_register(wb, data: "dd.DocumentData"):
    ws = wb.create_sheet("10. Source & Methodology")
    set_widths(ws, {1: 32, 2: 20, 3: 70})
    style_title(ws, 1, "Source & Methodology Register", span=3)
    r = 3
    r = note_block(ws, r, [
        "Every major output in this workbook traces to a canonical source or model record. Status key: CANONICAL =",
        "directly from data/canonical/*.yml. MODELLED = computed by tools/master_financial_model.py from canonical",
        "inputs. HISTORICAL-SUPERSEDED = retained for traceability only, no longer the basis for any live figure.",
        "BOUNDED = a range, not a point estimate. UNRESOLVED = no defensible figure exists yet.",
    ], span=3)
    r += 1
    headers = ["Output", "Status", "Source Chain"]
    style_header_row(ws, r, headers); r += 1

    for i, trace in enumerate(data.traceability):
        chain = " -> ".join(trace["chain"])
        write_row(ws, r, [trace["output"], "MODELLED", chain], alt=(i % 2 == 0), bold_col1=True)
        ws.row_dimensions[r].height = 45
        ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    r += 1

    style_section(ws, r, "Additional Key Outputs — Manually Mapped", span=3); r += 1
    style_header_row(ws, r, headers); r += 1
    extra = [
        ("Table 1/2 Steady-State Revenue", "CANONICAL", "data/canonical/revenue_ramp.yml#ramp_table1_m5plus / #ramp_table2_m5plus"),
        ("Table 1/2 Steady-State Net Operating Result (post-super)", "MODELLED", "tools/master_financial_model.py#compute_month_pnl -> data/canonical/cost_ramp.yml (incl. 2026-08-09 superannuation fix, item 46)"),
        ("Historical A$157,792.16 / A$118,297.16 revenue", "HISTORICAL-SUPERSEDED", "data/canonical/revenue_assumptions.yml#rev_historical_table1_monthly_inherited / _table2 — preserved, not deleted, not used in any live figure"),
        ("Historical A$63,028.75 (Table 1 Monthly Net P&L)", "HISTORICAL — NOT REVENUE", "docs/CURRENT-STATE.md (pre-2026-08-09-superannuation-fix figure) — this is a historical Net P&L figure, never revenue, superseded by this workbook's Sheet 4"),
        ("Bounded Funding Requirement (A$357,390-577,180)", "BOUNDED", "data/models/master_financial_model.yml#funding_requirement_investigation.combined_funding_requirement_bounded.primary_method"),
        ("Operating Cash Trough (both scenarios)", "MODELLED", "tools/master_financial_model.py#compute_cash_flow — accrual-basis proxy, disclosed simplification"),
        ("Break-Even AM Client Volume (both scenarios)", "MODELLED", "tools/master_financial_model.py#compute_breakeven — AM-volume-only method, disclosed scope limitation"),
        ("Startup-capital 6-9-range reconciliation", "UNRESOLVED", "docs/architecture/STARTUP-COST-RECONCILIATION.md — unchanged by this workbook"),
    ]
    for i, (label, status, chain) in enumerate(extra):
        write_row(ws, r, [label, status, chain], alt=(i % 2 == 0), bold_col1=True)
        ws.row_dimensions[r].height = 30
        ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    ws.freeze_panes = "A4"
    return ws


# ---------------------------------------------------------------------------
# Assembly
# ---------------------------------------------------------------------------
def main():
    data = dd.get_data()
    wb = Workbook()
    # Sheet 1 (build_executive_summary) reuses the workbook's default active sheet.

    build_executive_summary(wb, data)
    build_revenue_model(wb, data)
    build_cost_model(wb, data)
    build_pnl_model(wb, data)
    build_cash_flow(wb, data)
    build_breakeven(wb, data)
    build_funding_requirement(wb, data)
    build_sensitivity(wb, data)
    build_assumptions_open_issues(wb, data)
    build_source_register(wb, data)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Excel workbook written: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
